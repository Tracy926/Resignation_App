import argparse
import os
from dataclasses import dataclass
from datetime import date, datetime
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils.datetime import from_excel
import tkinter as tk
from tkinter import filedialog, messagebox

try:
    from plyer import notification as plyer_notification
except Exception:
    plyer_notification = None

try:
    from win10toast import ToastNotifier
except Exception:
    ToastNotifier = None


SHEET_NAME = "Resigned Employee List"


def normalize_header(value: str) -> str:
    return "".join(ch.lower() for ch in str(value) if ch.isalnum())


HEADER_ALIASES = {
    "EmpNo": ["empno", "employeeid", "employeecode"],
    "resignation_date": ["resignationdate"],
    # Combined header patterns (old format)
    "deactivation_system_combined": [
        normalize_header("Batch Deactivation from UM and 3rd Party Systems/Apps"),
        normalize_header("Batch Deactivation from UM & 3rd Party Systems/Apps"),
        normalize_header("Batch Deactivation from UM and Third Party Systems/Apps"),
    ],
    "deactivation_all_combined": [
        normalize_header(
            "Batch Deactivation from UM, 3rd Party Systems/Apps, E-mails, and Windows"
        ),
        normalize_header(
            "Batch Deactivation from UM, 3rd Party Systems/Apps, Emails, and Windows"
        ),
        normalize_header(
            "Batch Deactivation from UM, Third Party Systems/Apps, E-mails, and Windows"
        ),
    ],
    # Split headers (new format)
    "deactivation_um": [
        normalize_header("Batch Deactivation from UM"),
        normalize_header("UM"),
    ],
    "deactivation_3rd_party": [
        normalize_header("3rd Party Systems/Apps"),
        normalize_header("Third Party Systems/Apps"),
    ],
    "deactivation_emails": [
        normalize_header("Emails"),
        normalize_header("E-mails"),
    ],
    "deactivation_windows": [
        normalize_header("Windows"),
    ],
}

REQUIRED_KEYS = ["EmpNo", "resignation_date"]


@dataclass
class EmployeeRow:
    row_index: int
    EmpNo: str
    resignation_date: date
    can_system: bool
    can_all: bool


class DesktopNotifier:
    def __init__(self):
        self._toaster = ToastNotifier() if ToastNotifier else None

    def notify(self, title: str, message: str, timeout: int = 7):
        if plyer_notification:
            try:
                plyer_notification.notify(
                    title=title,
                    message=message,
                    app_name="Resigned Employee Deactivation",
                    timeout=timeout,
                )
                return
            except Exception:
                pass

        if self._toaster:
            try:
                self._toaster.show_toast(
                    title,
                    message,
                    duration=timeout,
                    threaded=False,
                )
                return
            except Exception:
                pass

        print(f"[{title}] {message}")


class DeactivationApp:
    def __init__(self, workbook_path: str, forced_output_path: Optional[str] = None):
        self.workbook_path = workbook_path
        self.forced_output_path = forced_output_path
        self.notifier = DesktopNotifier()

        # pandas read with openpyxl backend as requested.
        self.preview_df = pd.read_excel(
            workbook_path,
            sheet_name=SHEET_NAME,
            engine="openpyxl",
            header=None,
        )

        self.workbook = openpyxl.load_workbook(workbook_path)

        if SHEET_NAME not in self.workbook.sheetnames:
            raise ValueError(f"Worksheet '{SHEET_NAME}' not found in file.")

        self.sheet = self.workbook[SHEET_NAME]
        self.columns, self.header_row = self._find_required_columns()

        self.today = date.today()
        self.rows_to_process = self._find_due_rows()

        self.root = tk.Tk()
        self.root.title("Resigned Employee Batch Deactivation")
        self.root.geometry("900x500")

        self.selection_by_row: Dict[int, str] = {}

    def _has_content(self, value) -> bool:
        if value is None:
            return False
        if isinstance(value, str):
            return bool(value.strip())
        return True

    def _action_target_keys(self, action: str) -> List[str]:
        if action == "system":
            keys: List[str] = []
            if self.columns.get("deactivation_system_combined"):
                keys.append("deactivation_system_combined")
            else:
                keys.extend(["deactivation_um", "deactivation_3rd_party"])
            keys.extend(["deactivation_emails", "deactivation_windows"])
            return keys

        if self.columns.get("deactivation_all_combined"):
            return ["deactivation_all_combined", "deactivation_system_combined"]
        return [
            "deactivation_um",
            "deactivation_3rd_party",
            "deactivation_emails",
            "deactivation_windows",
        ]

    def _can_apply_action(self, row_index: int, action: str) -> bool:
        for key in self._action_target_keys(action):
            col = self.columns.get(key)
            if not col:
                continue
            value = self.sheet.cell(row=row_index, column=col).value
            if not self._has_content(value):
                return True
        return False

    def _infer_date_number_format(self, column: int) -> str:
        fallback_format = None
        for row in range(self.header_row + 1, self.sheet.max_row + 1):
            cell = self.sheet.cell(row=row, column=column)
            if not self._has_content(cell.value):
                continue

            number_format = cell.number_format
            if not number_format or str(number_format).strip().lower() == "general":
                continue

            if isinstance(cell.value, (date, datetime)):
                return number_format

            if fallback_format is None:
                fallback_format = number_format

        if fallback_format:
            return fallback_format

        return "m/d/yyyy"

    def _is_close_match(self, key: str, normalized_header: str) -> bool:
        if key == "EmpNo":
            return "emp" in normalized_header and (
                "no" in normalized_header or "id" in normalized_header
            )

        if key == "resignation_date":
            return "resign" in normalized_header and "date" in normalized_header

        if key == "deactivation_system_combined":
            return (
                "batch" in normalized_header
                and "deactivation" in normalized_header
                and "um" in normalized_header
                and "3rd" in normalized_header
                and "party" in normalized_header
                and "system" in normalized_header
            )

        if key == "deactivation_all_combined":
            return (
                "batch" in normalized_header
                and "deactivation" in normalized_header
                and "um" in normalized_header
                and "3rd" in normalized_header
                and "party" in normalized_header
                and "email" in normalized_header
                and "window" in normalized_header
            )

        if key == "deactivation_um":
            return "um" in normalized_header and "deactivation" in normalized_header

        if key == "deactivation_3rd_party":
            return (
                "3rd" in normalized_header
                and "party" in normalized_header
                and "system" in normalized_header
            )

        if key == "deactivation_emails":
            return "email" in normalized_header

        if key == "deactivation_windows":
            return "window" in normalized_header

        return False

    def _score_column_for_key(self, key: str, normalized_values: Set[str]) -> int:
        if not normalized_values:
            return 0

        score = 0
        aliases = HEADER_ALIASES.get(key, [])

        for value in normalized_values:
            if value in aliases:
                score += 5
            elif self._is_close_match(key, value):
                score += 2

        return score

    def _find_required_columns(self) -> Tuple[Dict[str, int], int]:
        best_normalized_to_index: Dict[str, int] = {}
        best_score = -1
        best_header_row = 1
        max_header_row_scan = min(200, self.sheet.max_row)

        for header_row in range(1, max_header_row_scan + 1):
            row_cells = list(
                self.sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=False)
            )[0]
            normalized_to_index: Dict[str, int] = {}

            for cell in row_cells:
                if cell.value is None:
                    continue
                normalized = normalize_header(str(cell.value))
                if normalized:
                    normalized_to_index[normalized] = cell.column

            score = 0
            for aliases in HEADER_ALIASES.values():
                if any(alias in normalized_to_index for alias in aliases):
                    score += 1

            if score > best_score:
                best_score = score
                best_normalized_to_index = normalized_to_index
                best_header_row = header_row

            if score == len(HEADER_ALIASES):
                break

        index_by_key: Dict[str, int] = {}
        missing_keys: List[str] = []

        for key, aliases in HEADER_ALIASES.items():
            matched_col = None
            for alias in aliases:
                if alias in best_normalized_to_index:
                    matched_col = best_normalized_to_index[alias]
                    break
            if matched_col is None:
                missing_keys.append(key)
            else:
                index_by_key[key] = matched_col

        if missing_keys:
            unresolved: List[str] = []
            used_columns = set(index_by_key.values())
            for key in missing_keys:
                fallback_col = None
                for normalized_header, column_idx in best_normalized_to_index.items():
                    if column_idx in used_columns:
                        continue
                    if self._is_close_match(key, normalized_header):
                        fallback_col = column_idx
                        break

                if fallback_col is None:
                    unresolved.append(key)
                else:
                    index_by_key[key] = fallback_col
                    used_columns.add(fallback_col)

            missing_keys = unresolved

        if missing_keys:
            max_rows_scan = min(500, self.sheet.max_row)
            max_cols_scan = self.sheet.max_column
            values_by_column: Dict[int, Set[str]] = {
                col: set() for col in range(1, max_cols_scan + 1)
            }

            for row in range(1, max_rows_scan + 1):
                for col in range(1, max_cols_scan + 1):
                    value = self.sheet.cell(row=row, column=col).value
                    if value is None:
                        continue
                    if not isinstance(value, str):
                        continue
                    normalized = normalize_header(value)
                    if normalized:
                        values_by_column[col].add(normalized)

            unresolved = []
            used_columns = set(index_by_key.values())

            for key in missing_keys:
                best_col = None
                best_col_score = 0

                for col, col_values in values_by_column.items():
                    if col in used_columns:
                        continue

                    col_score = self._score_column_for_key(key, col_values)
                    if col_score > best_col_score:
                        best_col_score = col_score
                        best_col = col

                if best_col is None or best_col_score == 0:
                    unresolved.append(key)
                else:
                    index_by_key[key] = best_col
                    used_columns.add(best_col)

            missing_keys = unresolved

        if missing_keys:
            found_headers = sorted(best_normalized_to_index.keys())
            sample_headers_by_col: List[str] = []
            max_rows_scan = min(100, self.sheet.max_row)
            for col in range(1, self.sheet.max_column + 1):
                for row in range(1, max_rows_scan + 1):
                    value = self.sheet.cell(row=row, column=col).value
                    if isinstance(value, str) and value.strip():
                        sample_headers_by_col.append(f"C{col}:{normalize_header(value)}")
                        break

            missing_required = [key for key in missing_keys if key in REQUIRED_KEYS]

            if missing_required:
                raise ValueError(
                    "Required column(s) not found: "
                    + ", ".join(missing_required)
                    + f". Best header row: {best_header_row}"
                    + ". Found normalized headers: "
                    + ", ".join(found_headers)
                    + ". Column samples: "
                    + ", ".join(sample_headers_by_col)
                )

        return index_by_key, best_header_row

    def _set_today_if_present(self, key: str, row_index: int):
        col = self.columns.get(key)
        if col:
            cell = self.sheet.cell(row=row_index, column=col)
            if not self._has_content(cell.value):
                cell.value = self.today
                cell.number_format = self._infer_date_number_format(col)

    def _set_text_if_present(self, key: str, row_index: int, text: str):
        col = self.columns.get(key)
        if col:
            cell = self.sheet.cell(row=row_index, column=col)
            if not self._has_content(cell.value):
                cell.value = text

    def _apply_system(self, row_index: int):
        if self.columns.get("deactivation_system_combined"):
            self._set_today_if_present("deactivation_system_combined", row_index)
        else:
            self._set_today_if_present("deactivation_um", row_index)
            self._set_today_if_present("deactivation_3rd_party", row_index)

        self._set_text_if_present("deactivation_emails", row_index, "No existing account")
        self._set_text_if_present("deactivation_windows", row_index, "No existing account")

        self.selection_by_row[row_index] = "system"

    def _apply_all(self, row_index: int):
        if self.columns.get("deactivation_all_combined"):
            self._set_today_if_present("deactivation_all_combined", row_index)
            self._set_today_if_present("deactivation_system_combined", row_index)
        else:
            self._set_today_if_present("deactivation_um", row_index)
            self._set_today_if_present("deactivation_3rd_party", row_index)
            self._set_today_if_present("deactivation_emails", row_index)
            self._set_today_if_present("deactivation_windows", row_index)

        self.selection_by_row[row_index] = "all"

    def _find_due_rows(self) -> List[EmployeeRow]:
        due: List[EmployeeRow] = []
        emp_col = self.columns["EmpNo"]
        resign_col = self.columns["resignation_date"]

        for row in range(self.header_row + 1, self.sheet.max_row + 1):
            resignation_date = self._excel_value_to_date(
                self.sheet.cell(row=row, column=resign_col).value
            )
            if resignation_date is None:
                continue
            if resignation_date < self.today:
                emp_no_val = self.sheet.cell(row=row, column=emp_col).value
                can_system = self._can_apply_action(row, "system")
                can_all = self._can_apply_action(row, "all")

                if not can_system and not can_all:
                    continue

                due.append(
                    EmployeeRow(
                        row_index=row,
                        EmpNo=str(emp_no_val) if emp_no_val is not None else "",
                        resignation_date=resignation_date,
                        can_system=can_system,
                        can_all=can_all,
                    )
                )

        return due

    def _standardize_target_columns_format(self):
        target_keys = [
            "deactivation_system_combined",
            "deactivation_all_combined",
            "deactivation_um",
            "deactivation_3rd_party",
            "deactivation_emails",
            "deactivation_windows",
        ]

        for key in target_keys:
            col = self.columns.get(key)
            if not col:
                continue

            for row in range(self.header_row + 1, self.sheet.max_row + 1):
                cell = self.sheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Arial", size=10)

    def _excel_value_to_date(self, value) -> Optional[date]:
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, (int, float)):
            try:
                converted = from_excel(value, self.workbook.epoch)
                if isinstance(converted, datetime):
                    return converted.date()
                if isinstance(converted, date):
                    return converted
            except Exception:
                return None

        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None

            patterns = [
                "%Y-%m-%d",
                "%d/%m/%Y",
                "%m/%d/%Y",
                "%d-%m-%Y",
                "%m-%d-%Y",
            ]
            for pattern in patterns:
                try:
                    return datetime.strptime(text, pattern).date()
                except ValueError:
                    continue

        return None

    def _refresh_status(self):
        processed = len(self.selection_by_row)
        total = len(self.rows_to_process)
        self.status_var.set(f"Processed: {processed}/{total}")
        self.save_button.config(
            state=tk.NORMAL if processed == total and total > 0 else tk.DISABLED
        )

    def _build_ui(self):
        title = tk.Label(
            self.root,
            text="Employees with Resignation Date Before Today",
            font=("Segoe UI", 12, "bold"),
        )
        title.pack(pady=10)

        if not self.rows_to_process:
            msg = tk.Label(
                self.root,
                text="No employees found with resignation date before today.",
                font=("Segoe UI", 10),
            )
            msg.pack(pady=20)
            self.notifier.notify(
                "Resigned Employee Deactivation",
                "No employees found with resignation date before today.",
            )

            close_btn = tk.Button(
                self.root, text="Close", command=self.root.destroy, width=14
            )
            close_btn.pack(pady=10)
            return

        self.notifier.notify(
            "Resigned Employee Deactivation",
            f"{len(self.rows_to_process)} employees are ready for action.",
        )

        container = tk.Frame(self.root)
        container.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        canvas = tk.Canvas(container)
        scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")),
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        header = tk.Frame(scrollable_frame)
        header.pack(fill=tk.X, pady=(0, 6))
        tk.Label(
            header,
            text="EmpNo",
            width=22,
            anchor="w",
            font=("Segoe UI", 10, "bold"),
        ).pack(side="left")
        tk.Label(
            header,
            text="Resignation Date",
            width=18,
            anchor="w",
            font=("Segoe UI", 10, "bold"),
        ).pack(side="left")
        tk.Label(
            header,
            text="Actions",
            anchor="w",
            font=("Segoe UI", 10, "bold"),
        ).pack(side="left")

        for item in self.rows_to_process:
            row_frame = tk.Frame(scrollable_frame)
            row_frame.pack(fill=tk.X, pady=2)

            tk.Label(row_frame, text=item.EmpNo, width=22, anchor="w").pack(side="left")
            tk.Label(
                row_frame,
                text=item.resignation_date.isoformat(),
                width=18,
                anchor="w",
            ).pack(side="left")

            btn_holder = tk.Frame(row_frame)
            btn_holder.pack(side="left")

            status_lbl = tk.Label(btn_holder, text="Pending", width=10, fg="darkorange")
            status_lbl.pack(side="right", padx=6)

            def update_row_visuals(
                selected: Optional[str],
                system_btn: tk.Button,
                all_btn: tk.Button,
                status_label: tk.Label,
            ):
                if selected == "system":
                    status_label.config(text="System", fg="green")
                    system_btn.config(relief=tk.SUNKEN)
                    all_btn.config(relief=tk.RAISED)
                elif selected == "all":
                    status_label.config(text="All", fg="green")
                    system_btn.config(relief=tk.RAISED)
                    all_btn.config(relief=tk.SUNKEN)
                else:
                    status_label.config(text="Pending", fg="darkorange")
                    system_btn.config(relief=tk.RAISED)
                    all_btn.config(relief=tk.RAISED)

            def on_select(
                choice: str,
                row_index: int,
                system_btn: tk.Button,
                all_btn: tk.Button,
                status_label: tk.Label,
            ):
                current_choice = self.selection_by_row.get(row_index)

                if current_choice == choice:
                    self.selection_by_row.pop(row_index, None)
                    update_row_visuals(None, system_btn, all_btn, status_label)
                else:
                    self.selection_by_row[row_index] = choice
                    update_row_visuals(choice, system_btn, all_btn, status_label)

                self._refresh_status()

            system_btn = tk.Button(btn_holder, text="System", width=12)
            all_btn = tk.Button(btn_holder, text="All", width=12)
            system_btn.pack(side="left", padx=3)
            all_btn.pack(side="left", padx=3)

            if not item.can_system:
                system_btn.config(state=tk.DISABLED)
            if not item.can_all:
                all_btn.config(state=tk.DISABLED)

            system_btn.config(
                command=lambda r=item.row_index, s=system_btn, a=all_btn, l=status_lbl: on_select(
                    "system", r, s, a, l
                )
            )
            all_btn.config(
                command=lambda r=item.row_index, s=system_btn, a=all_btn, l=status_lbl: on_select(
                    "all", r, s, a, l
                )
            )

        footer = tk.Frame(self.root)
        footer.pack(fill=tk.X, pady=10, padx=12)

        self.status_var = tk.StringVar(value="Processed: 0/0")
        status = tk.Label(footer, textvariable=self.status_var, anchor="w")
        status.pack(side="left")

        self.save_button = tk.Button(
            footer,
            text="Save Updated File",
            state=tk.DISABLED,
            command=self._save_output,
        )
        self.save_button.pack(side="right")

        cancel_button = tk.Button(footer, text="Cancel", command=self.root.destroy)
        cancel_button.pack(side="right", padx=6)

        self._refresh_status()

    def _save_output(self):
        if self.forced_output_path:
            output_path = self.forced_output_path
        else:
            today_str = datetime.now().strftime("%Y%m%d")
            default_name = (
                os.path.splitext(os.path.basename(self.workbook_path))[0]
                + f"_updated_{today_str}.xlsx"
            )
            initial_dir = os.path.dirname(self.workbook_path)

            output_path = filedialog.asksaveasfilename(
                title="Save Updated Excel File",
                initialdir=initial_dir,
                initialfile=default_name,
                defaultextension=".xlsx",
                filetypes=[("Excel Workbook", "*.xlsx")],
            )

        if not output_path:
            return

        for item in self.rows_to_process:
            choice = self.selection_by_row.get(item.row_index)
            if choice == "system":
                self._apply_system(item.row_index)
            elif choice == "all":
                self._apply_all(item.row_index)

        self._standardize_target_columns_format()
        self.workbook.save(output_path)

        messagebox.showinfo("Success", f"Updated file saved to:\n{output_path}")
        self.notifier.notify(
            "Resigned Employee Deactivation",
            f"Saved updated file: {os.path.basename(output_path)}",
            timeout=10,
        )
        self.root.destroy()

    def run(self):
        self._build_ui()
        self.root.mainloop()


def choose_input_file() -> Optional[str]:
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(
        title="Select Input Excel File",
        filetypes=[("Excel Files", "*.xlsx;*.xlsm")],
    )

    root.destroy()
    return file_path if file_path else None


def main():
    parser = argparse.ArgumentParser(
        description="Update deactivation columns for resigned employees with popup selection UI."
    )
    parser.add_argument(
        "--input",
        help="Path to the input Excel file. If omitted, default/download picker is used.",
    )
    parser.add_argument(
        "--output",
        help="Optional fixed output path. If omitted, Save dialog is shown.",
    )
    args = parser.parse_args()

    input_file = args.input
    if not input_file:
        default_input = os.path.join(
            os.path.expanduser("~"),
            "Downloads",
            "Resignations_sample.xlsx",
        )
        if os.path.exists(default_input):
            input_file = default_input
            print(f"No --input provided. Using default: {input_file}")
        else:
            input_file = choose_input_file()

    if not input_file:
        print("No input file selected.")
        return

    if not os.path.exists(input_file):
        print(f"Input file not found: {input_file}")
        return

    try:
        app = DeactivationApp(input_file, forced_output_path=args.output)
        app.run()
    except Exception as exc:
        DesktopNotifier().notify("Resigned Employee Deactivation - Error", str(exc), timeout=10)
        messagebox.showerror("Error", str(exc))


if __name__ == "__main__":
    main()
